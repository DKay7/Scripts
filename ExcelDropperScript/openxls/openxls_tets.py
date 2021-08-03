import openpyxl
from time import time
import numpy as np
from openpyxl import Workbook

start1 = time()
wb_read = openpyxl.load_workbook('../ковры.xlsx', read_only=True)
print(f"read: {time() - start1}")

wb_write = Workbook(write_only=True)

ws_read = wb_read[wb_read.sheetnames[0]]
ws_write = wb_write.create_sheet()

start = time()
list_to_stay = ['id', 'url', "group_id"]
storage_read = np.array(tuple(ws_read.rows), dtype=object)
print(storage_read.shape)
storage_read.reshape((ws_read.max_row, ws_read.max_column))
wb_read.close()
del wb_read
print(storage_read.shape)
storage_write = np.zeros(shape=(storage_read.shape[0], len(list_to_stay)), dtype=object)


column = 0
for col in storage_read.transpose():

    if col[0].value in list_to_stay:
        row = 0
        for cell in col:
            storage_write[row][column] = cell.value
            row += 1
        column += 1

for row in storage_write:
    ws_write.append(row.tolist())

print(f"Finish: {time() - start}")

start = time()
wb_write.save('../ковры_new_test1.xlsx')
print(f"Write: {time() - start}")
print(f"TOTAL: {time() - start1}")
